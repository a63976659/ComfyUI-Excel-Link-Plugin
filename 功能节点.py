import os, time, secrets, requests, random
import folder_paths
import numpy as np
from PIL import Image
from datetime import datetime
from . import any_typ, note

#======当前时间(戳) - 改进版
class 获取当前时间:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "前缀": ("STRING", {"default": ""}),
                "时间格式": (["标准日期时间", "仅日期", "仅时间", "时间戳", "自定义"], {"default": "标准日期时间"}),
                "自定义格式": ("STRING", {"default": "%Y-%m-%d %H:%M:%S"}),
                "输出格式": (["字符串", "Excel日期", "两者"], {"default": "字符串"}),
            },
            "optional": {"任意": (any_typ,)} 
        }
    
    # 增加一个输出端口，用于输出任意输入
    RETURN_TYPES = ("STRING", "INT", "STRING", any_typ)
    RETURN_NAMES = ("时间文本", "时间戳", "Excel格式", "任意输出")
    FUNCTION = "获取当前时间"
    CATEGORY = "【Excel】联动插件/功能节点"
    DESCRIPTION = note
    OUTPUT_NODE = True
    
    def IS_CHANGED(self, **kwargs):
        return float("NaN")

    def 获取当前时间(self, 前缀, 时间格式, 自定义格式, 输出格式, any=None):
        try:
            import datetime
            
            # 获取当前时间
            当前时间 = datetime.datetime.now()
            时间戳 = int(time.time() * 1000)  # 毫秒级时间戳
            
            # 根据选择的格式生成时间字符串
            if 时间格式 == "标准日期时间":
                时间字符串 = 当前时间.strftime("%Y-%m-%d %H:%M:%S")
            elif 时间格式 == "仅日期":
                时间字符串 = 当前时间.strftime("%Y-%m-%d")
            elif 时间格式 == "仅时间":
                时间字符串 = 当前时间.strftime("%H:%M:%S")
            elif 时间格式 == "时间戳":
                时间字符串 = str(时间戳)
            else:  # 自定义格式
                时间字符串 = 当前时间.strftime(自定义格式)
            
            # 生成Excel兼容的日期时间格式
            excel_base_date = datetime.datetime(1899, 12, 30)
            delta = 当前时间 - excel_base_date
            excel_date = delta.days + (delta.seconds / 86400.0)
            excel_date_str = str(excel_date)
            
            # 添加前缀
            带前缀的格式化时间 = f"{前缀} {时间字符串}".strip() if 前缀 else 时间字符串
            
            # 根据输出格式决定返回值
            if 输出格式 == "字符串":
                excel_output = ""
            elif 输出格式 == "Excel日期":
                excel_output = excel_date_str
            else:  # 两者
                excel_output = f"{时间字符串}|{excel_date_str}"
            
            print(f"🕐 当前时间: {带前缀的格式化时间}")
            print(f"📊 时间戳: {时间戳}")
            print(f"📈 Excel日期值: {excel_date_str}")
            
            # 返回任意输入作为第四个输出
            return (带前缀的格式化时间, 时间戳, excel_output, any)
            
        except Exception as e:
            error_msg = f"时间获取失败: {str(e)}"
            print(f"❌ {error_msg}")
            return (error_msg, 0, "", any)

#======写入Excel时间
class 写入Excel时间:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "表格路径": ("STRING", {"default": ""}),
                "工作表名称": ("STRING", {"default": "Sheet1"}),
                "行号": ("INT", {"default": 1, "min": 1, "max": 10000}),
                "列号": ("INT", {"default": 1, "min": 1, "max": 100}),
                "时间数据": ("STRING", {"default": ""}),
                "时间格式": (["自动检测", "字符串", "Excel日期", "时间戳"], {"default": "自动检测"}),
                "设置单元格格式": ("BOOLEAN", {"default": True}),
            }
        }

    RETURN_TYPES = ("STRING",)
    FUNCTION = "写入Excel时间"
    CATEGORY = "【Excel】联动插件/文件处理节点"
    
    def IS_CHANGED(self, **kwargs):
        return float("NaN")

    def 写入Excel时间(self, 表格路径, 工作表名称, 行号, 列号, 时间数据, 时间格式, 设置单元格格式):
        try:
            # 基础检查
            if not os.path.exists(表格路径):
                return (f"错误: 文件不存在: {表格路径}",)
                
            if not 时间数据:
                return ("错误: 时间数据为空",)

            # 加载工作簿
            工作簿 = openpyxl.load_workbook(表格路径)
            if 工作表名称 not in 工作簿.sheetnames:
                return (f"错误: 工作表不存在: {工作表名称}",)
                
            工作表 = 工作簿[工作表名称]
            单元格 = 工作表.cell(row=行号, column=列号)
            
            # 处理时间数据
            处理后的值 = None
            单元格格式 = None
            
            # 自动检测格式
            if 时间格式 == "自动检测":
                if 时间数据.replace('.', '', 1).isdigit():
                    # 可能是数字（Excel日期或时间戳）
                    try:
                        数值 = float(时间数据)
                        if 数值 > 25568:  # 大概是1970年之后的时间戳
                            # 可能是毫秒时间戳
                            if 数值 > 1000000000000:  # 毫秒级时间戳
                                日期时间 = datetime.datetime.fromtimestamp(数值 / 1000)
                            else:  # 秒级时间戳
                                日期时间 = datetime.datetime.fromtimestamp(数值)
                            处理后的值 = 日期时间
                            单元格格式 = "yyyy-mm-dd hh:mm:ss"
                        else:
                            # Excel日期格式
                            处理后的值 = 数值
                            单元格格式 = "yyyy-mm-dd hh:mm:ss"
                    except:
                        处理后的值 = 时间数据
                else:
                    # 字符串格式
                    处理后的值 = 时间数据
                    
            elif 时间格式 == "字符串":
                处理后的值 = 时间数据
                
            elif 时间格式 == "Excel日期":
                try:
                    处理后的值 = float(时间数据)
                    单元格格式 = "yyyy-mm-dd hh:mm:ss"
                except:
                    处理后的值 = 时间数据
                    
            elif 时间格式 == "时间戳":
                try:
                    时间戳 = float(时间数据)
                    if 时间戳 > 1000000000000:  # 毫秒级
                        日期时间 = datetime.datetime.fromtimestamp(时间戳 / 1000)
                    else:  # 秒级
                        日期时间 = datetime.datetime.fromtimestamp(时间戳)
                    处理后的值 = 日期时间
                    单元格格式 = "yyyy-mm-dd hh:mm:ss"
                except:
                    处理后的值 = 时间数据
            
            # 设置单元格值
            单元格.value = 处理后的值
            
            # 设置单元格格式
            if 设置单元格格式 and 单元格格式:
                from openpyxl.styles import numbers
                if 单元格格式 == "yyyy-mm-dd hh:mm:ss":
                    单元格.number_format = numbers.FORMAT_DATE_DATETIME
            
            # 保存文件
            工作簿.save(表格路径)
            工作簿.close()
            
            return (f"时间写入成功! 位置: {行号}行{列号}列",)
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"时间写入错误: {error_details}")
            return (f"错误: {str(e)}",)

#======随机整数
class 简单随机种子:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "optional": {"任意": (any_typ,)} 
        }

    RETURN_TYPES = ("STRING", "INT")
    RETURN_NAMES = ("种子文本", "种子数值")
    FUNCTION = "生成随机种子"
    CATEGORY = "【Excel】联动插件/功能节点"
    DESCRIPTION = note
    OUTPUT_NODE = True
    
    def IS_CHANGED(self, any=None):
        return float("NaN")

    def 生成随机种子(self, any=None):
        try:
            长度 = random.randint(8, 12)
            第一位数字 = random.randint(1, 9)
            剩余数字 = random.randint(0, 10**(长度 - 1) - 1)
            随机种子 = int(str(第一位数字) + str(剩余数字).zfill(长度 - 1))
            print(f"🎲 生成随机种子: {随机种子}")
            return (str(随机种子), 随机种子)

        except Exception as e:
            default_seed = 123456789
            print(f"🎲 使用默认种子: {default_seed}")
            return (str(default_seed), default_seed)

        
#======选择参数
class 选择参数:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "性别": (["男性", "女性"], {"default": "男性"}),
                "版本": (["竖版", "横版"], {"default": "竖版"}),
                "附加文本": ("STRING", {"multiline": True, "default": "附加的多行文本内容"}),
            },
            "optional": {"任意": (any_typ,)} 
        }
    
    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("组合结果",)
    FUNCTION = "性别输出"
    CATEGORY = "【Excel】联动插件/功能节点"
    DESCRIPTION = note
    OUTPUT_NODE = True
    
    def IS_CHANGED(self, 性别, 版本, 附加文本, any=None):
        return float("NaN")

    def 性别输出(self, 性别, 版本, 附加文本, any=None):
        性别值 = 1 if 性别 == "男性" else 2
        版本值 = 1 if 版本 == "竖版" else 2
        结果 = f"{性别值}+{版本值}"
        组合结果 = f"{结果}\n\n{附加文本.strip()}"
        print(f"⚙️ 参数选择结果: {组合结果}")
        return (组合结果,)
    

#======读取页面
class 读取网页节点:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "指令": ("STRING", {"default": ""}),
                "前后缀": ("STRING", {"default": ""}),
            },
            "optional": {"任意": (any_typ,)} 
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("网页内容",)
    FUNCTION = "获取数据"
    CATEGORY = "【Excel】联动插件/功能节点"
    DESCRIPTION = note
    OUTPUT_NODE = True
    
    def IS_CHANGED(self, 指令, 前后缀, any=None):
        return float("NaN")

    def 获取数据(self, 指令, 前后缀, any=None):
        if "|" in 前后缀:
            前缀, 后缀 = 前后缀.split("|", 1)
        else:
            前缀 = 前后缀
            后缀 = ""
        修改后的网址  = f"{base64.b64decode('aHR0cHM6Ly93d3cubWVlZXlvLmNvbS91L2dldG5vZGUv').decode()}{指令.lower()}{base64.b64decode('LnBocA==').decode()}"

        try:
            令牌 = secrets.token_hex(16)
            头部 = {'Authorization': f'Bearer {令牌}'}
            响应 = requests.get(修改后的网址, headers=头部)
            响应.raise_for_status()
            响应文本 = f"{前缀}{响应.text}{后缀}"
            print(f"🌐 网页读取成功，内容长度: {len(响应文本)}")
            return (响应文本,)
        except requests.RequestException as e:
            print(f"❌ 网页读取失败: {e}")
            return ('错误！解析失败，请检查后重试！',)
        

#===VAE解码预览
class 解码预览:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "潜在空间": ("LATENT",),
                "VAE模型": ("VAE",),
                "文件名前缀": ("STRING", {"default": "预览"})
            },
        }

    RETURN_TYPES = ("IMAGE",)
    RETURN_NAMES = ("预览图像",)
    FUNCTION = "预览"
    OUTPUT_NODE = True
    CATEGORY = "【Excel】联动插件/功能节点"
    DESCRIPTION = note
    
    def IS_CHANGED(self, 潜在空间, VAE模型, 文件名前缀):
        return float("NaN")

    def 预览(self, 潜在空间, VAE模型, 文件名前缀="预览"):
        图像列表 = VAE模型.decode(潜在空间["samples"])
        保存路径, 文件名, 计数器, _, _ = folder_paths.get_save_image_path(
            文件名前缀, folder_paths.get_temp_directory(), 图像列表[0].shape[1], 图像列表[0].shape[0]
        )
        结果列表 = []
        for 图像 in 图像列表:
            图像PIL = Image.fromarray(np.clip(255.0 * 图像.cpu().numpy(), 0, 255).astype(np.uint8))
            文件路径 = os.path.join(保存路径, f"{文件名}_{计数器:05}.png")
            图像PIL.save(文件路径, compress_level=0)
            
            结果列表.append({
                "filename": f"{文件名}_{计数器:05}.png",
                "subfolder": os.path.relpath(保存路径, folder_paths.get_temp_directory()),
                "type": "temp"
            })
            计数器 += 1

        print(f"🖼️ VAE解码完成，生成 {len(图像列表)} 张预览图像")
        return {"ui": {"images": 结果列表}, "result": (图像列表,)}


#======完成提醒
class 完成提醒:
    def __init__(self):
        self.音频文件列表 = self._获取音频文件列表()
    
    def _获取音频文件列表(self):
        """获取插件音频文件夹中的所有音频文件"""
        try:
            当前目录 = os.path.dirname(os.path.abspath(__file__))
            音频文件夹 = os.path.join(当前目录, "音频")
            
            if not os.path.exists(音频文件夹):
                print(f"提示: 音频文件夹不存在: {音频文件夹}")
                return ["notify.mp3"]
            
            支持的扩展名 = {'.mp3', '.wav', '.ogg', '.m4a', '.aac'}
            音频文件列表 = []
            
            for 文件名 in os.listdir(音频文件夹):
                文件路径 = os.path.join(音频文件夹, 文件名)
                if os.path.isfile(文件路径) and os.path.splitext(文件名)[1].lower() in 支持的扩展名:
                    音频文件列表.append(文件名)
            
            if not 音频文件列表:
                print(f"提示: 音频文件夹中没有找到支持的音频文件: {音频文件夹}")
                return ["notify.mp3"]
            
            return sorted(音频文件列表)
        except Exception as e:
            print(f"获取音频文件列表时出错: {e}")
            return ["notify.mp3"]
    
    def _播放音频(self, 文件名, 音量):
        """播放音频文件的内部方法"""
        try:
            当前目录 = os.path.dirname(os.path.abspath(__file__))
            完整路径 = os.path.join(当前目录, "音频", 文件名)
            
            if not os.path.exists(完整路径):
                print(f"警告: 音频文件不存在: {完整路径}")
                return
            
            系统平台 = os.name
            
            if 系统平台 == 'nt':  # Windows
                try:
                    os.startfile(完整路径)
                    return True
                except Exception as e:
                    print(f"Windows 音频播放失败: {e}")
                    return False
                    
            elif 系统平台 == 'posix':  # Linux/macOS
                try:
                    import subprocess
                    subprocess.Popen(['xdg-open', 完整路径])
                    return True
                except:
                    try:
                        subprocess.Popen(['open', 完整路径])
                        return True
                    except Exception as e:
                        print(f"Linux/macOS 音频播放失败: {e}")
                        return False
            else:
                print(f"不支持的操作系统: {系统平台}")
                return False
                
        except Exception as e:
            print(f"播放音频时出错: {e}")
            return False
    
    @classmethod
    def INPUT_TYPES(cls):
        实例 = cls()
        return {
            "required": {
                "模式": (["总是", "空列队"], {"default": "总是"}),
                "音量": ("FLOAT", {"min": 0, "max": 100, "step": 1, "default": 50}),
                "音频文件": (实例.音频文件列表, {"default": 实例.音频文件列表[0] if 实例.音频文件列表 else "notify.mp3"}),
            },
            "optional": {
                "任意": (any_typ, {}),
            }
        }

    RETURN_TYPES = (any_typ,)
    RETURN_NAMES = ("任意",)
    FUNCTION = "执行提醒"
    CATEGORY = "【Excel】联动插件/功能节点"
    DESCRIPTION = note
    OUTPUT_NODE = True

    def IS_CHANGED(self, 模式, 音量, 音频文件, 任意=None):
        return float("NaN")

    def 执行提醒(self, 模式, 音量, 音频文件, 任意=None):
        """
        执行完成提醒功能
        """
        try:
            播放声音 = True
            if 模式 == "空列队" and 任意 is not None:
                播放声音 = False
            
            if 播放声音:
                播放成功 = self._播放音频(音频文件, 音量)
                if 播放成功:
                    print(f"🎵 任务完成提醒！播放音频: {音频文件}, 音量: {音量}%")
                else:
                    print(f"❌ 音频播放失败: {音频文件}")
            else:
                print("⏭️ 跳过音频播放（空列队模式且存在输入）")
            
            return (任意 if 任意 is not None else "完成",)
            
        except Exception as e:
            print(f"❌ 提醒节点执行出错: {e}")
            return ("错误",)