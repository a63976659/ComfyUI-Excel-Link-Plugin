import os, re, io, base64, shutil, torch, openpyxl
import numpy as np
from PIL import Image as PILImage
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
import folder_paths
from . import any_typ, note

#====== 替换文件名
class 替换文件名:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "文件路径": ("STRING", {"default": "path/to/your/file.jpg"}),
                "新文件名": ("STRING", {"default": ""}),
            },
            "optional": {"任意": (any_typ,)} 
        }

    RETURN_TYPES = ("STRING",)
    FUNCTION = "执行替换"
    CATEGORY = "【Excel】联动插件/文件处理节点"
    DESCRIPTION = "修改指定路径文件的名称，但保留其原始后缀名。常用于批量生成后对文件进行重命名管理。"
    
    def 执行替换(self, 文件路径, 新文件名, any=None):
        dir_name = os.path.dirname(文件路径)
        _, ext = os.path.splitext(文件路径)
        clean_name = re.sub(r'[\/:*?"<>|]', '_', 新文件名)
        new_path = os.path.join(dir_name, clean_name + ext)
        return (new_path,)

#====== 文件路径和后缀统计
class 文件路径和后缀统计:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "文件夹路径": ("STRING", {"default": ""}),
                "文件扩展名": (["jpg", "png", "jpg&png", "txt", "csv", "全部"], {"default": "jpg"}),
            },
            "optional": {"任意": (any_typ,)} 
        }

    RETURN_TYPES = ("STRING", "INT", "LIST")
    RETURN_NAMES = ("路径列表文本", "文件数量", "路径列表对象")
    FUNCTION = "统计文件"
    CATEGORY = "【Excel】联动插件/文件处理节点"
    DESCRIPTION = "扫描指定文件夹，根据后缀名过滤文件。返回合并的路径文本、文件总数以及一个可供循环使用的路径列表。"

    def 统计文件(self, 文件夹路径, 文件扩展名, any=None):
        if not os.path.isdir(文件夹路径): return ("", 0, [])
        exts = ('.jpg', '.png') if 文件扩展名 == "jpg&png" else (f".{文件扩展名}" if 文件扩展名 != "全部" else "")
        files = [os.path.join(文件夹路径, f) for f in os.listdir(文件夹路径) 
                 if os.path.isfile(os.path.join(文件夹路径, f)) and f.lower().endswith(exts)]
        return ('\n'.join(files), len(files), files)

#====== 图像层叠加
class 图像层叠加:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "底层图像": ("IMAGE",),
                "上层图像": ("IMAGE",),
                "对齐方式": (["居中", "左上", "右上", "左下", "右下"], {"default": "居中"}),
                "缩放比例": ("FLOAT", {"default": 1.0, "min": 0.1, "max": 10.0, "step": 0.1}),
                "不透明度": ("FLOAT", {"default": 1.0, "min": 0.0, "max": 1.0, "step": 0.1}),
            }
        }

    RETURN_TYPES = ("IMAGE",)
    FUNCTION = "执行叠加"
    CATEGORY = "【Excel】联动插件/文件处理节点"
    DESCRIPTION = "将两张图像进行叠加处理。支持调整上层图的缩放比例、不透明度以及相对于底层图的对齐位置。"

    def 执行叠加(self, 底层图像, 上层图像, 对齐方式, 缩放比例, 不透明度):
        bg = PILImage.fromarray((底层图像[0].cpu().numpy() * 255).astype(np.uint8)).convert("RGBA")
        fg = PILImage.fromarray((上层图像[0].cpu().numpy() * 255).astype(np.uint8)).convert("RGBA")
        
        fg = fg.resize((int(fg.width * 缩放比例), int(fg.height * 缩放比例)), PILImage.LANCZOS)
        if 不透明度 < 1.0:
            alpha = fg.split()[3].point(lambda p: p * 不透明度)
            fg.putalpha(alpha)
            
        x, y = 0, 0
        if 对齐方式 == "居中":
            x, y = (bg.width - fg.width) // 2, (bg.height - fg.height) // 2
        elif 对齐方式 == "右上": x = bg.width - fg.width
        elif 对齐方式 == "左下": y = bg.height - fg.height
        elif 对齐方式 == "右下": x, y = bg.width - fg.width, bg.height - fg.height

        bg.paste(fg, (x, y), fg)
        res = torch.from_numpy(np.array(bg.convert("RGB")).astype(np.float32) / 255.0).unsqueeze(0)
        return (res,)

#====== 读取表格数据
class 读取Excel数据:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "表格路径": ("STRING", {"default": ""}),
                "工作表名称": ("STRING", {"default": "Sheet1"}),
                "行范围": ("STRING", {"default": "2-2"}),
                "列范围": ("STRING", {"default": "1-1"}),
            }
        }
    RETURN_TYPES = ("STRING",)
    FUNCTION = "执行读取"
    CATEGORY = "【Excel】联动插件/文件处理节点"
    DESCRIPTION = "从Excel中提取文字。支持单行单列或范围读取。多个数据将以'|'分隔。"

    def 执行读取(self, 表格路径, 工作表名称, 行范围, 列范围):
        try:
            wb = openpyxl.load_workbook(表格路径, data_only=True, read_only=True)
            ws = wb[工作表名称]
            r_start, r_end = (map(int, 行范围.split('-')) if '-' in 行范围 else (int(行范围), int(行范围)))
            c_start, c_end = (map(int, 列范围.split('-')) if '-' in 列范围 else (int(列范围), int(列范围)))
            
            lines = []
            for r in range(r_start, r_end + 1):
                row_vals = [str(ws.cell(r, c).value or "") for c in range(c_start, c_end + 1)]
                lines.append("|".join(row_vals))
            return ("\n".join(lines),)
        except Exception as e:
            return (f"读取失败: {str(e)}",)

#====== 写入表格数据
class 写入Excel数据:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "表格路径": ("STRING", {"default": ""}),
                "工作表名称": ("STRING", {"default": "Sheet1"}),
                "起始行": ("INT", {"default": 2, "min": 1}),
                "起始列": ("INT", {"default": 1, "min": 1}),
                "数据内容": ("STRING", {"multiline": True, "default": ""}),
            }
        }
    RETURN_TYPES = ("STRING",)
    FUNCTION = "执行写入"
    CATEGORY = "【Excel】联动插件/文件处理节点"
    DESCRIPTION = "将文本写入指定位置。如果输入多行文本，程序会自动换行写入。"

    def 执行写入(self, 表格路径, 工作表名称, 起始行, 起始列, 数据内容):
        try:
            wb = openpyxl.load_workbook(表格路径)
            ws = wb[工作表名称]
            for i, line in enumerate(数据内容.split('\n')):
                for j, val in enumerate(line.split('|')):
                    ws.cell(row=起始行 + i, column=起始列 + j).value = val.strip()
            wb.save(表格路径)
            return ("写入成功",)
        except Exception as e:
            return (f"写入失败: {str(e)}",)

#====== 图片插入表格
class 写入Excel图片:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "表格路径": ("STRING", {"default": ""}),
                "工作表名称": ("STRING", {"default": "Sheet1"}),
                "行范围": ("STRING", {"default": "1"}),
                "列范围": ("STRING", {"default": "1"}),
                "图片路径": ("STRING", {"default": ""}),
                "缩放模式": (["匹配单元格", "固定尺寸", "原图大小"], {"default": "匹配单元格"}),
                "图片宽度": ("INT", {"default": 300}),
                "图片高度": ("INT", {"default": 200}),
                "跨行数": ("INT", {"default": 1, "min": 1}),
                "跨列数": ("INT", {"default": 1, "min": 1}),
            }
        }
    RETURN_TYPES = ("STRING",)
    FUNCTION = "执行插入"
    CATEGORY = "【Excel】联动插件/文件处理节点"
    DESCRIPTION = "将图像插入Excel单元格。选择'匹配单元格'时图片根据行列宽自动调整。"

    def 执行插入(self, 表格路径, 工作表名称, 行范围, 列范围, 图片路径, 缩放模式, 图片宽度, 图片高度, 跨行数, 跨列数):
        try:
            row = int(re.split(r'[-;；,，\s]+', 行范围.strip())[0])
            col = int(re.split(r'[-;；,，\s]+', 列范围.strip())[0])
            
            wb = openpyxl.load_workbook(表格路径)
            ws = wb[工作表名称]
            
            with PILImage.open(图片路径) as img:
                if 缩放模式 == "匹配单元格":
                    w = (ws.column_dimensions[get_column_letter(col)].width or 10) * 7 * 跨列数
                    h = (ws.row_dimensions[row].height or 15) * 1.33 * 跨行数
                    img = img.resize((int(w), int(h)), PILImage.LANCZOS)
                elif 缩放模式 == "固定尺寸":
                    img = img.resize((图片宽度, 图片高度), PILImage.LANCZOS)

                img_stream = io.BytesIO()
                img.convert("RGB").save(img_stream, format="PNG")
                img_stream.seek(0)
                ox_img = OpenpyxlImage(img_stream)
                ws.add_image(ox_img, get_column_letter(col) + str(row))
            
            wb.save(表格路径)
            return (f"图片已插入至 {row}行{col}列",)
        except Exception as e:
            return (f"插入失败: {str(e)}",)

#====== 查找表格数据
class 查找Excel数据:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "表格路径": ("STRING", {"default": ""}),
                "工作表名称": ("STRING", {"default": "Sheet1"}),
                "查找内容": ("STRING", {"default": ""}),
                "查找模式": (["精确查找", "模糊查找"], {"default": "精确查找"}),
            }
        }
    RETURN_TYPES = ("STRING", "INT", "INT")
    RETURN_NAMES = ("结果文本", "行号", "列号")
    FUNCTION = "执行查找"
    CATEGORY = "【Excel】联动插件/文件处理节点"
    DESCRIPTION = "在工作表中搜索指定内容，返回其所在的行号和列号。"

    def 执行查找(self, 表格路径, 工作表名称, 查找内容, 查找模式):
        try:
            wb = openpyxl.load_workbook(表格路径, data_only=True, read_only=True)
            ws = wb[工作表名称]
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    val = str(ws.cell(r, c).value or "")
                    if (查找模式 == "精确查找" and val == 查找内容) or (查找模式 == "模糊查找" and 查找内容 in val):
                        return (f"找到: {r}行{c}列", r, c)
            return ("未找到", 0, 0)
        except Exception as e:
            return (f"查找错误: {str(e)}", 0, 0)

#====== 读取表格数量统计/差值
class 读取Excel行列差:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "表格路径": ("STRING", {"default": ""}),
                "工作表名称": ("STRING", {"default": "Sheet1"}),
                "读取模式": (["读行", "读列"], {"default": "读行"}),
                "索引": ("STRING", {"default": "1,3"}),
            }
        }
    RETURN_TYPES = ("INT",)
    FUNCTION = "计算逻辑"
    CATEGORY = "【Excel】联动插件/文件处理节点"
    DESCRIPTION = "输入单值(如 1)统计非空总数；输入多值(如 1,3)计算数量差。支持中英文及空格分隔。"

    def 计算逻辑(self, 表格路径, 工作表名称, 读取模式, 索引):
        try:
            wb = openpyxl.load_workbook(表格路径, data_only=True, read_only=True)
            ws = wb[工作表名称]
            
            # 统一解析中英文分隔符
            parts = re.split(r'[,，;；\s|]+', 索引.strip())
            indices = [int(p) for p in parts if p.strip().isdigit()]
            
            if not indices: return (0,)

            def count_non_empty(idx):
                count = 0
                if 读取模式 == "读行":
                    for c in range(1, ws.max_column + 1):
                        if ws.cell(idx, c).value is not None: count += 1
                else:
                    for r in range(1, ws.max_row + 1):
                        if ws.cell(r, idx).value is not None: count += 1
                return count

            if len(indices) == 1:
                return (count_non_empty(indices[0]),)
            else:
                return (count_non_empty(indices[0]) - count_non_empty(indices[1]),)
                
        except Exception as e:
            print(f"统计失败: {str(e)}")
            return (0,)

#====== 写入Excel时间
class 写入Excel时间:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "表格路径": ("STRING", {"default": ""}),
                "工作表名称": ("STRING", {"default": "Sheet1"}),
                "行范围": ("STRING", {"default": "1"}),
                "列范围": ("STRING", {"default": "1"}),
                "时间数据": ("STRING", {"default": ""}),
            }
        }
    RETURN_TYPES = ("STRING",)
    FUNCTION = "执行写入"
    CATEGORY = "【Excel】联动插件/文件处理节点"
    DESCRIPTION = "向指定单元格写入时间文本。支持灵活的范围格式解析。"

    def 执行写入(self, 表格路径, 工作表名称, 行范围, 列范围, 时间数据):
        try:
            row = int(re.split(r'[-;；,，\s]+', 行范围.strip())[0])
            col = int(re.split(r'[-;；,，\s]+', 列范围.strip())[0])
            
            wb = openpyxl.load_workbook(表格路径)
            ws = wb[工作表名称]
            ws.cell(row=row, column=col).value = 时间数据
            wb.save(表格路径)
            return (f"时间已写入至 {row}行{col}列",)
        except Exception as e:
            return (f"写入失败: {str(e)}",)